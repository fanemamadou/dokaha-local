from django.shortcuts import render, redirect
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, Q, F
from django.contrib.auth.models import User
from django.contrib import messages
from decimal import Decimal
from gestion.models import ProductionOeufs, Vente, Depense, Poulailler, CategorieStock, MouvementStock, Produit
from .forms import StockEntreeForm, StockSortieForm

# =============================================================================
# 📦 VUES STOCK (placeholders garantis fonctionnels)
# =============================================================================
@login_required
def stock_historique(request):
    """Historique des mouvements Stock"""
    return render(request, 'gestion/base.html', {'title': '📜 Historique Stock', 'user': request.user})

# =============================================================================
# 🎯 AUTRES VUES (placeholders si absentes)
# =============================================================================

@login_required
def dashboard(request):
    """Dashboard: KPIs + Actions + Activité Récente"""
    from .models import Poulailler, Vente, Depense, SortiePoules
    from django.db.models import Sum
    from django.utils import timezone
    from datetime import timedelta
    from django.shortcuts import render

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    total_cheptel = Poulailler.objects.aggregate(total=Sum('effectif_initial'))['total'] or 0
    mortalite_today = SortiePoules.objects.filter(date=today, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0
    mortalite_week = SortiePoules.objects.filter(date__gte=week_ago, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0

    try:
        total_recettes = Vente.objects.aggregate(total=Sum('montant_total'))['total'] or 0
        total_depenses = Depense.objects.aggregate(total=Sum('montant'))['total'] or 0
    except:
        total_recettes = 0
        try: total_depenses = Depense.objects.aggregate(total=Sum('montant_total'))['total'] or 0
        except: total_depenses = 0

    tresorerie = total_recettes - total_depenses
    alerte_mortalite = mortalite_week > (total_cheptel * 0.02) if total_cheptel > 0 else False
    alerte_tresorerie = tresorerie < 0

    # 📜 Activité Récente (mix Ventes + Mortalités)
    activites = []
    try:
        for v in Vente.objects.order_by('-date')[:3]:
            activites.append({'date': v.date, 'type': '💰 Vente', 'detail': f"{v.type_vente} - {v.montant_total:,.0f} F", 'info': v.client or 'N/A'})
        for m in SortiePoules.objects.filter(type_sortie='mortalite').order_by('-date')[:2]:
            activites.append({'date': m.date, 'type': '💀 Mortalité', 'detail': f"{m.nombre} sujets", 'info': 'Cheptel'})
        activites.sort(key=lambda x: x['date'], reverse=True)
        activites = activites[:5]
    except:
        activites = []

    return render(request, 'gestion/dashboard.html', {
        'total_cheptel': total_cheptel, 'mortalite_today': mortalite_today,
        'mortalite_week': mortalite_week, 'tresorerie': tresorerie,
        'alerte_mortalite': alerte_mortalite, 'alerte_tresorerie': alerte_tresorerie,
        'activites_recentes': activites
    })


@login_required
def dg_dashboard(request):
    return render(request, 'gestion/base.html', {'title': 'Dashboard DG', 'user': request.user})

@login_required
def vente_list(request):
    ventes = Vente.objects.all().order_by('-date')[:100]
    return render(request, 'gestion/vente_list.html', {'ventes': ventes, 'title': 'Ventes', 'user': request.user})

@login_required
def vente_choice(request):
    return render(request, 'gestion/vente_choice.html', {'title': 'Type de vente', 'user': request.user})


@login_required
def vente_form(request, type_vente):
    """Formulaire vente unifié : oeufs/poulets/autres + déduction sécurisée"""
    from .models import Vente, Poulailler
    from django.contrib import messages
    from django.utils import timezone
    from django.shortcuts import redirect
    from django.db import transaction

    poulaillers = Poulailler.objects.all()

    if request.method == 'POST':
        client = request.POST.get('client', '').strip()
        unites = float(request.POST.get('unites', 0) or 0)
        prix = float(request.POST.get('prix', 0) or 0)
        montant_paye = float(request.POST.get('montant_paye', 0) or 0)
        mode_paiement = request.POST.get('mode_paiement', 'espece')
        poulailler_id = request.POST.get('poulailler_id')
        designation = request.POST.get('designation', '').strip()

        if not client or unites <= 0 or prix <= 0:
            messages.error(request, "❌ Client, quantité et prix obligatoires.")
        elif type_vente == 'autres' and not designation:
            messages.error(request, "❌ Désignation obligatoire pour ce type.")
        else:
            try:
                with transaction.atomic():
                    montant_total = unites * prix
                    montant_restant = max(0, montant_total - montant_paye)
                    paye = montant_restant == 0
                    msg = ""

                    # 🐔 Déduction poulailler pour poulets
                    if type_vente == 'poulets' and poulailler_id:
                        p = Poulailler.objects.select_for_update().get(id=poulailler_id)
                        if p.effectif_initial < unites:
                            raise ValueError(f"⚠️ Effectif insuffisant dans {p.nom} ({p.effectif_initial} dispo)")
                        p.effectif_initial -= int(unites)
                        p.save()
                        msg = f"🐔 Déduit de {p.nom} | "

                    Vente.objects.create(
                        date=timezone.now().date(), client=client, type_vente=type_vente,
                        unites=unites, prix_plateau=prix, montant_total=montant_total,
                        montant_paye=montant_paye, montant_restant=montant_restant,
                        paye=paye, mode_paiement=mode_paiement,
                        poulailler_id=poulailler_id if type_vente == 'poulets' else None
                    )
                    messages.success(request, f"✅ Vente validée ! {msg}Total: {montant_total:,.0f} FCFA")
                    return redirect('dashboard')
            except Poulailler.DoesNotExist:
                messages.error(request, "❌ Poulailler introuvable.")
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"❌ Erreur: {e}")

    return render(request, 'gestion/vente_form.html', {
        'type_vente': type_vente, 'poulaillers': poulaillers, 'now': timezone.now()
    })


@login_required
def acces_refuse(request):
    return render(request, 'gestion/acces_refuse.html', {'title': '🚫 Accès Refusé', 'user': request.user})

# 📦 VUES STOCK - Ajoutées proprement (ne pas toucher)
@login_required
def stock_historique(request):
    return render(request, 'gestion/base.html', {'title': '📜 Historique Stock', 'user': request.user})
@login_required
@login_required
def stock_ajouter(request):
    """Formulaire mouvement stock (compatible catégorie + choices réels)"""
    from .forms import StockMouvementForm
    from .models import MouvementStock, CategorieStock
    from django.utils import timezone
    from django.contrib import messages
    from decimal import Decimal
    
    if request.method == 'POST':
        form = StockMouvementForm(request.POST)
        if form.is_valid():
            cat = form.cleaned_data['categorie']  # Le form renvoie maintenant une CategorieStock
            type_mvt = form.cleaned_data['type_mouvement']  # 'entree', 'sortie', etc.
            quantite = Decimal(str(form.cleaned_data['quantite']))
            commentaire = form.cleaned_data.get('commentaire', '')
            
            MouvementStock.objects.create(
                categorie=cat,
                date=timezone.now(),
                type_mouvement=type_mvt,
                quantite=quantite,
                prix_unitaire=Decimal('0'),
                montant_total=Decimal('0'),
                commentaire=commentaire,
                cree_par=request.user
            )
            messages.success(request, f"✅ {quantite} {cat.unite_mesure if cat.unite_mesure else 'unité(s)'} {type_mvt} enregistrée(s)")
            return redirect('/stock/')
    else:
        form = StockMouvementForm()
        
    return render(request, 'gestion/stock_ajouter.html', {
        'title': '📝 Ajouter un Mouvement Stock',
        'form': form,
        'user': request.user
    })
@login_required
def stock_dashboard(request):
    """Dashboard Stock V2 : indicateurs visuels + ratios"""
    from django.db.models import Sum
    from gestion.models import CategorieStock, MouvementStock

    categories = CategorieStock.objects.all().order_by('nom')
    data = []
    total_alertes = 0
    
    for cat in categories:
        entrees = MouvementStock.objects.filter(categorie=cat, type_mouvement='entree').aggregate(Sum('quantite'))['quantite__sum'] or 0
        sorties = MouvementStock.objects.filter(categorie=cat, type_mouvement='sortie').aggregate(Sum('quantite'))['quantite__sum'] or 0
        ajust   = MouvementStock.objects.filter(categorie=cat, type_mouvement='ajustement').aggregate(Sum('quantite'))['quantite__sum'] or 0
        pertes  = MouvementStock.objects.filter(categorie=cat, type_mouvement='perte').aggregate(Sum('quantite'))['quantite__sum'] or 0
        
        stock = float(entrees - sorties + ajust - pertes)
        seuil = cat.seuil_alerte or 0
        is_critical = seuil > 0 and stock <= seuil
        if is_critical: total_alertes += 1
        
        # Ratio pour barre de progression (max 100%)
        ratio = min(100, round((stock / seuil * 100) if seuil > 0 else 0, 1))
        
        data.append({
            'cat': cat, 'stock': stock, 'seuil': seuil,
            'is_critical': is_critical, 'ratio': ratio
        })
        
    return render(request, 'gestion/stock_dashboard.html', {
        'title': '📦 État des Stocks',
        'categories': data,
        'total_categories': len(data),
        'total_alertes': total_alertes,
        'user': request.user
    })

# 🥚 VUE COLLECTE ŒUFS (DUAL MODE + SYNC STOCK)







@login_required
def collecte_oeufs(request):
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from django.utils import timezone
    from datetime import datetime
    from .models import Poulailler, Collecte
    poulaillers = Poulailler.objects.all()
    if request.method == 'POST':
        try:
            pid=request.POST.get('poulailler_id')
            ds=request.POST.get('date')
            dc=datetime.strptime(ds,'%Y-%m-%d').date() if ds else timezone.localdate()
            plt=int(request.POST.get('plateaux',0) or 0)
            vrc=int(request.POST.get('oeufs_vrac',0) or 0)
            cas=int(request.POST.get('oeufs_casses',0) or 0)
            tot=(plt*30)+vrc
            if not pid:
                messages.error(request,'❌ Sélectionne un poulailler.')
                return redirect('collecte_oeufs')
            Collecte.objects.create(date=dc,poulailler_id=pid,plateaux=plt,oeufs_unites=vrc,total_oeufs=tot,oeufs_casses=cas)
            messages.success(request,'✅ Collecte validée : '+str(tot)+' œufs sains ('+str(plt)+' plt + '+str(vrc)+' vrac) + '+str(cas)+' cassés')
            return redirect('terrain')
        except Exception as e:
            messages.error(request,'❌ Erreur: '+str(e))
            return redirect('collecte_oeufs')
    return render(request,'gestion/collecte_form.html',{'poulaillers':poulaillers,'today':timezone.now()})


@login_required
def mortalite_form(request):
    from .forms import MortaliteForm
    from .models import Poulailler, SortiePoules
    from django.db.models import Sum
    from django.utils import timezone
    from django.contrib import messages

    # Préparer les données du tableau
    poulaillers_data = []
    for p in Poulailler.objects.all():
        m = SortiePoules.objects.filter(poulailler=p, type_sortie='mortalite').aggregate(t=Sum('nombre'))['t'] or 0
        poulaillers_data.append({'nom': p.nom, 'initial': p.effectif_initial, 'mortalites': m, 'reste': p.effectif_initial - m})

    if request.method == 'POST':
        form = MortaliteForm(request.POST)
        if form.is_valid():
            p = form.cleaned_data['poulailler']
            n = form.cleaned_data['nombre']
            if n <= p.effectif_initial:
                p.effectif_initial -= n
                p.save()
                SortiePoules.objects.create(date=form.cleaned_data['date'], poulailler=p, nombre=n, type_sortie='mortalite')
                messages.success(request, f"✅ {n} poules déduites de '{p.nom}'")
                return redirect('mortalite_form')
            else:
                messages.error(request, f"⚠️ Effectif insuffisant ({p.effectif_initial} < {n})")
    else:
        form = MortaliteForm(initial={'date': timezone.now().date()})

    return render(request, 'gestion/mortalite_form.html', {'form': form, 'poulaillers': poulaillers_data})


@login_required
def dashboard(request):
    """Dashboard: KPIs + Actions + Activité Récente"""
    from .models import Poulailler, Vente, Depense, SortiePoules
    from django.db.models import Sum
    from django.utils import timezone
    from datetime import timedelta
    from django.shortcuts import render

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    total_cheptel = Poulailler.objects.aggregate(total=Sum('effectif_initial'))['total'] or 0
    mortalite_today = SortiePoules.objects.filter(date=today, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0
    mortalite_week = SortiePoules.objects.filter(date__gte=week_ago, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0

    try:
        total_recettes = Vente.objects.aggregate(total=Sum('montant_total'))['total'] or 0
        total_depenses = Depense.objects.aggregate(total=Sum('montant'))['total'] or 0
    except:
        total_recettes = 0
        try: total_depenses = Depense.objects.aggregate(total=Sum('montant_total'))['total'] or 0
        except: total_depenses = 0

    tresorerie = total_recettes - total_depenses
    alerte_mortalite = mortalite_week > (total_cheptel * 0.02) if total_cheptel > 0 else False
    alerte_tresorerie = tresorerie < 0

    # 📜 Activité Récente (mix Ventes + Mortalités)
    activites = []
    try:
        for v in Vente.objects.order_by('-date')[:3]:
            activites.append({'date': v.date, 'type': '💰 Vente', 'detail': f"{v.type_vente} - {v.montant_total:,.0f} F", 'info': v.client or 'N/A'})
        for m in SortiePoules.objects.filter(type_sortie='mortalite').order_by('-date')[:2]:
            activites.append({'date': m.date, 'type': '💀 Mortalité', 'detail': f"{m.nombre} sujets", 'info': 'Cheptel'})
        activites.sort(key=lambda x: x['date'], reverse=True)
        activites = activites[:5]
    except:
        activites = []

    return render(request, 'gestion/dashboard.html', {
        'total_cheptel': total_cheptel, 'mortalite_today': mortalite_today,
        'mortalite_week': mortalite_week, 'tresorerie': tresorerie,
        'alerte_mortalite': alerte_mortalite, 'alerte_tresorerie': alerte_tresorerie,
        'activites_recentes': activites
    })


@login_required
def credits_list(request):
    """Liste des créances clients impayées"""
    from .models import Vente
    from django.db.models import Sum
    from django.shortcuts import render
    creances = Vente.objects.filter(paye=False, montant_restant__gt=0).order_by('-date')
    total = creances.aggregate(total=Sum('montant_restant'))['total'] or 0
    return render(request, 'gestion/credits_list.html', {'creances': creances, 'total_credits': total})

@login_required
def credits_relance(request, vente_id):
    """Générer un message de relance pour un client"""
    from .models import Vente
    from django.shortcuts import get_object_or_404, render
    vente = get_object_or_404(Vente, pk=vente_id)
    msg = (f"Bonjour {vente.client},\n"
           f"Concernant votre achat du {vente.date.strftime('%d/%m/%Y')}:\n"
           f"📦 Total: {vente.montant_total:,.0f} FCFA | ✅ Déjà versé: {vente.montant_paye or 0:,.0f} FCFA\n"
           f"⚠️ *Reste à payer: {vente.montant_restant:,.0f} FCFA*\n\n"
           f"Merci de régulariser dans les meilleurs délais.\nCordialement, Dg Tropic 🇨🇮")
    return render(request, 'gestion/credits_relance.html', {'vente': vente, 'message': msg})

@login_required
def credits_payer(request, vente_id):
    """Marquer une créance comme réglée"""
    from .models import Vente
    from django.shortcuts import get_object_or_404, redirect
    from django.contrib import messages
    vente = get_object_or_404(Vente, pk=vente_id)
    vente.paye = True
    vente.montant_restant = 0
    vente.save()
    messages.success(request, f"✅ Créance de {vente.montant_total:,.0f} FCFA clôturée.")
    return redirect('credits_list')


@login_required
def credits_paiement_partiel(request, vente_id):
    """Gérer un paiement partiel ou total (avec Decimal pour éviter les erreurs)"""
    from .models import Vente
    from django.shortcuts import get_object_or_404, redirect, render
    from django.contrib import messages
    from decimal import Decimal

    vente = get_object_or_404(Vente, pk=vente_id)

    if request.method == 'POST':
        montant_str = request.POST.get('montant', '0')
        try:
            montant = Decimal(montant_str)
        except:
            messages.error(request, "❌ Montant invalide.")
            return redirect('credits_paiement', vente_id=vente_id)

        if montant <= 0:
            messages.error(request, "❌ Le montant doit être supérieur à 0.")
        elif montant > vente.montant_restant:
            messages.error(request, f"⚠️ Dépasse le reste dû ({vente.montant_restant:,.0f} F).")
        else:
            # ✅ Tous les calculs en Decimal
            vente.montant_paye = (vente.montant_paye or Decimal(0)) + montant
            vente.montant_restant = vente.montant_restant - montant
            
            if vente.montant_restant <= 0:
                vente.montant_restant = Decimal(0)
                vente.paye = True
            vente.save()
            
            msg = "✅ Créance totalement réglée !" if vente.paye else f"✅ Acompte de {int(montant)} F enregistré."
            messages.success(request, f"{msg} Reste: {int(vente.montant_restant)} FCFA")
            return redirect('credits_list')

    return render(request, 'gestion/credits_paiement.html', {'vente': vente})


@login_required
def finance_dashboard(request):
    from .models import Vente, Depense
    from django.db.models import Sum
    start, end = request.GET.get('start'), request.GET.get('end')
    qs_v, qs_d = Vente.objects.all(), Depense.objects.all()
    if start and end: qs_v, qs_d = qs_v.filter(date__range=[start, end]), qs_d.filter(date__range=[start, end])
    elif start: qs_v, qs_d = qs_v.filter(date__gte=start), qs_d.filter(date__gte=start)
    elif end: qs_v, qs_d = qs_v.filter(date__lte=end), qs_d.filter(date__lte=end)
    
    tr = int(qs_v.aggregate(t=Sum('montant_total'))['t'] or 0)
    td = int(qs_d.aggregate(t=Sum('montant'))['t'] or 0)
    
    labels, data = [], []
    for item in qs_v.values('type_vente').annotate(t=Sum('montant_total')).order_by('type_vente'):
        labels.append(item['type_vente']); data.append(int(item['t'] or 0))
        
    return render(request, 'gestion/finance_dashboard.html', {
        'total_recettes': tr, 'total_depenses': td, 'tresorerie': tr - td,
        'start_date': start or '', 'end_date': end or '',
        'chart_labels': labels, 'chart_data': data
    })

@login_required
def finance_export_excel(request):
    from .models import Vente, Depense
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws_r = wb.active; ws_r.title = "Recettes"
    ws_r.append(["Date", "Client", "Type", "Qté", "Prix", "Total", "Payé", "Reste"])
    for c in ws_r[1]: c.font = Font(bold=True); c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for v in Vente.objects.order_by('-date'):
        ws_r.append([v.date.strftime("%d/%m/%Y") if v.date else "", v.client or "", v.type_vente or "", int(v.unites or 0), int(v.prix_plateau or 0), int(v.montant_total or 0), int(v.montant_paye or 0), int(v.montant_restant or 0)])
    
    ws_d = wb.create_sheet("Dépenses")
    ws_d.append(["Date", "Catégorie", "Description", "Montant"])
    for c in ws_d[1]: c.font = Font(bold=True)
    for d in Depense.objects.order_by('-date'):
        ws_d.append([d.date.strftime("%d/%m/%Y") if d.date else "", d.categorie or "", d.description or "", int(d.montant or 0)])
        
    ws_rc = wb.create_sheet("Récap")
    ws_rc.append(["INDICATEUR", "VALEUR"]); ws_rc['A1'].font = Font(bold=True)
    tr = int(Vente.objects.aggregate(t=Sum('montant_total'))['t'] or 0)
    td = int(Depense.objects.aggregate(t=Sum('montant'))['t'] or 0)
    ws_rc.append(["Total Recettes", tr]); ws_rc.append(["Total Dépenses", td]); ws_rc.append(["Trésorerie", tr-td])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Tropic_Finance_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def cheptel(request):
    """Vue terrain : effectif réel par poulailler"""
    from .models import Poulailler, SortiePoules, Vente
    from django.db.models import Sum
    from django.shortcuts import render
    
    poulaillers = []
    for p in Poulailler.objects.all():
        effectif = p.effectif_initial or 0
        mortalite = SortiePoules.objects.filter(poulailler=p, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0
        ventes = Vente.objects.filter(poulailler=p, type_vente='poulets').aggregate(total=Sum('unites'))['total'] or 0
        effectif_reel = max(0, effectif - mortalite - int(ventes or 0))
        poulaillers.append({'nom': p.nom, 'initial': effectif, 'mortalite': int(mortalite or 0), 'vendus': int(ventes or 0), 'reel': effectif_reel, 'alerte': effectif_reel < (effectif * 0.8) if effectif > 0 else False})
    
    return render(request, 'gestion/cheptel.html', {'poulaillers': poulaillers, 'total_reel': sum(p['reel'] for p in poulaillers)})

@login_required
def terrain(request):
    """Interface terrain pour Flatie"""
    from .models import Poulailler, Vente, Depense, SortiePoules
    from django.db.models import Sum
    from django.utils import timezone
    from datetime import timedelta
    from django.shortcuts import render

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    total_cheptel = Poulailler.objects.aggregate(total=Sum('effectif_initial'))['total'] or 0
    mortalite_week = SortiePoules.objects.filter(date__gte=week_ago, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0
    
    try:
        tr = Vente.objects.aggregate(t=Sum('montant_total'))['t'] or 0
        td = Depense.objects.aggregate(t=Sum('montant'))['t'] or 0
    except: tr = td = 0
    
    alerte_mortalite = mortalite_week > (total_cheptel * 0.02) if total_cheptel > 0 else False
    alerte_tresorerie = (tr - td) < 0
    
    activites = []
    try:
        for v in Vente.objects.order_by('-date')[:3]:
            activites.append({'date': v.date, 'type': 'Vente', 'detail': f"{v.type_vente} - {int(v.montant_total or 0)} F"})
        for m in SortiePoules.objects.filter(type_sortie='mortalite').order_by('-date')[:2]:
            activites.append({'date': m.date, 'type': 'Mortalité', 'detail': f"{int(m.nombre or 0)} sujets"})
        activites.sort(key=lambda x: x['date'] or today, reverse=True)
    except: pass
    
    return render(request, 'gestion/terrain.html', {
        'total_cheptel': total_cheptel, 'alerte_mortalite': alerte_mortalite,
        'alerte_tresorerie': alerte_tresorerie, 'activites_recentes': activites[:5]
    })

@login_required
def cheptel(request):
    """Vue terrain : effectif réel par poulailler"""
    from .models import Poulailler, SortiePoules, Vente
    from django.db.models import Sum
    from django.shortcuts import render
    
    poulaillers = []
    for p in Poulailler.objects.all():
        effectif = p.effectif_initial or 0
        mort = SortiePoules.objects.filter(poulailler=p, type_sortie='mortalite').aggregate(total=Sum('nombre'))['total'] or 0
        vend = Vente.objects.filter(poulailler=p, type_vente='poulets').aggregate(total=Sum('unites'))['total'] or 0
        reel = max(0, effectif - mort - int(vend or 0))
        poulaillers.append({'nom': p.nom, 'initial': effectif, 'mortalite': int(mort or 0), 'vendus': int(vend or 0), 'reel': reel, 'alerte': reel < (effectif * 0.8) if effectif > 0 else False})
    
    return render(request, 'gestion/cheptel.html', {'poulaillers': poulaillers, 'total_reel': sum(p['reel'] for p in poulaillers)})

@login_required
def depense_add(request):
    """Formulaire terrain : dépenses générales, 1 poulailler ou partagées"""
    from .models import Depense, Poulailler
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from django.utils import timezone

    poulaillers = Poulailler.objects.all()
    categories = ['Aliment', 'Transport', "Main d'œuvre", 'Vaccin/Soin', 'Eau/Électricité', 'Matériel', 'Autre']

    if request.method == 'POST':
        type_dep = request.POST.get('type_depense', 'generale')
        categorie = request.POST.get('categorie', '').strip()
        montant_str = request.POST.get('montant', '').replace(' ', '').replace(',', '.')
        description = request.POST.get('description', '').strip()
        
        try:
            # Nettoyage format CI : on enlève espaces et virgules, on garde les chiffres
            clean_str = montant_str.replace(' ', '').replace(',', '')
            montant = float(clean_str) if clean_str else 0
        except ValueError:
            montant = 0
        except ValueError: montant = 0

        if not categorie or montant <= 0:
            messages.error(request, "❌ Catégorie et montant obligatoires.")
        else:
            # Construction de la description selon le type
            if type_dep == 'unique':
                p_id = request.POST.get('poulailler_id')
                p = Poulailler.objects.get(id=p_id) if p_id else None
                desc = f"[{p.nom}] {description}" if p else description
            elif type_dep == 'partagee':
                p_ids = request.POST.getlist('poulaillers_partages')
                noms = [Poulailler.objects.get(id=pid).nom for pid in p_ids]
                desc = f"[Partagée: {', '.join(noms)}] {description}"
            else:
                desc = f"[Générale] {description}"

            Depense.objects.create(
                date=timezone.now().date(),
                categorie=categorie,
                description=desc,
                montant=montant
            )
            messages.success(request, f"✅ Dépense de {int(montant)} F enregistrée.")
            return redirect('terrain')

    return render(request, 'gestion/depense_form.html', {
        'categories': categories,
        'poulaillers': poulaillers
    })

@login_required
def sante_form(request):
    from .models import Poulailler, SanteRecord
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from django.utils import timezone

    poulaillers = Poulailler.objects.all()
    if request.method == 'POST':
        type_dep = request.POST.get('type_depense', 'generale')
        type_sante = request.POST.get('type_sante', '').strip()
        produit = request.POST.get('produit', '').strip()
        cout_str = request.POST.get('cout', '').replace(' ', '').replace(',', '')
        note = request.POST.get('note', '').strip()
        try: cout = float(cout_str) if cout_str else 0
        except: cout = 0

        if not type_sante:
            messages.error(request, "❌ Choisis un type d'intervention.")
        else:
            p_obj = None
            if type_dep == 'unique':
                p_id = request.POST.get('poulailler_id')
                if p_id: p_obj = Poulailler.objects.get(id=p_id)
            elif type_dep == 'partagee':
                p_ids = request.POST.getlist('poulaillers_partages')
                noms = [Poulailler.objects.get(id=pid).nom for pid in p_ids]
                note = f"[Partagée: {', '.join(noms)}] {note}"

            SanteRecord.objects.create(date=timezone.now().date(), poulailler=p_obj, type_sante=type_sante, produit=produit, cout=cout, note=note)
            messages.success(request, f"✅ Santé enregistrée ({produit}).")
            return redirect('terrain')
    return render(request, 'gestion/sante_form.html', {'poulaillers': poulaillers})


@login_required
def rapport_jour(request):
    """Rapport quotidien terrain - Version sécurisée"""
    from .models import Poulailler, SortiePoules
    from django.db.models import Sum
    from django.shortcuts import render
    from datetime import date

    today = date.today()
    date_str = today.strftime("%d/%m/%Y")
    
    data = []
    t_mort, t_plt, t_oeufs = 0, 0, 0

    for p in Poulailler.objects.all():
        # Mortalité (toujours disponible)
        mort = int(SortiePoules.objects.filter(date=today, poulailler=p, type_sortie='mortalite').aggregate(t=Sum('nombre'))['t'] or 0)
        t_mort += mort
        
        # Collecte : tentative sécurisée avec fallback
        plt, oeufs = 0, 0
        try:
            from .models import Collecte
            c = Collecte.objects.filter(date=today, poulailler=p).first()
            if c:
                # Essaye plusieurs noms de champs possibles
                plt = int(getattr(c, 'plateaux', getattr(c, 'unites', 0)) or 0)
                oeufs = int(getattr(c, 'oeufs_unites', getattr(c, 'oeufs', getattr(c, 'unites', 0))) or 0)
        except:
            pass  # Si Collecte n'existe pas, on reste à 0
            
        t_plt += plt
        t_oeufs += oeufs
        
        # Traitement & Obs (sécurisé)
        try:
            from .models import SanteRecord
            treated = "Oui" if SanteRecord.objects.filter(date=today, poulailler=p).exists() else "Non"
            obs = ""
            rec = SanteRecord.objects.filter(date=today, poulailler=p).last()
            if rec and rec.note: obs = rec.note[:20]
        except:
            treated, obs = "Non", ""
            
        data.append({'nom': p.nom, 'mort': mort, 'plt': plt, 'oeufs': oeufs, 'tx': treated, 'obs': obs})

    # Texte WhatsApp
    txt = f" RAPPORT {date_str}\n"
    for d in data:
        txt += f" {d['nom']}: {d['plt']}plt+{d['oeufs']} | {d['mort']}m | {d['tx']}\n"
    txt += f" TOTAL: {t_plt}plt + {t_oeufs} | {t_mort}morts"

    return render(request, 'gestion/rapport_jour.html', {
        'date_str': date_str, 'data': data,
        't_mort': t_mort, 't_plt': t_plt, 't_oeufs': t_oeufs,
        'rapport_text': txt
    })
